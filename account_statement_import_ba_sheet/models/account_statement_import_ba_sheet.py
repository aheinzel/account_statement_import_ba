
import base64
import io
import logging
import re
from datetime import datetime, date
from typing import Dict, Optional, List, Tuple, Set

from odoo import _, models, fields
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# ---- Strict header lists (case-insensitive, exact labels only; no synonyms) ----
REQUIRED_HEADERS = [
    "operation date",
    "value date",
    "booking text",
    "internal note",
    "currency",
    "amount",
]

OPTIONAL_HEADERS = [
    "record data",
    "record number",
    "payer name",
    "payer account",
    "payer bank code",
    "payee name",
    "payee account",
    "payee bank code",
    "purpose text",
    "reference",
]

ALL_HEADERS = REQUIRED_HEADERS + OPTIONAL_HEADERS

def _norm(s: str) -> str:
    s = (s or "").strip()
    s = " ".join(s.split()).lower()
    return s

def _excel_kind(content: bytes) -> Optional[str]:
    sig_xlsx = bytes([0x50, 0x4B, 0x03, 0x04])  # ZIP 'PK..'
    sig_xls  = bytes([0xD0, 0xCF, 0x11, 0xE0, 0xA1, 0xB1, 0x1A, 0xE1])  # OLE CFBF
    if content[:4] == sig_xlsx:
        return "xlsx"
    if content[:8] == sig_xls:
        return "xls"
    return None

_DATE_PATTERNS = ["%Y-%m-%d","%d.%m.%Y","%d.%m.%y","%Y/%m/%d","%d/%m/%Y","%m/%d/%Y"]

def _to_iso_date(val) -> str:
    """Best-effort to get YYYY-MM-DD string (for logs and payload)."""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = (str(val or "").strip())
    for fmt in _DATE_PATTERNS:
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s).date().isoformat()
    except Exception:
        return s

def _parse_number(val) -> float:
    if val in (None, ""):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("\xa0"," ")
    s = s.replace(" ", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".","").replace(",",".")
        else:
            s = s.replace(",","")
    elif "," in s:
        s = s.replace(".","").replace(",",".")
    try:
        return float(s)
    except Exception:
        s2 = re.sub(r"[^0-9\.-]","", s)
        return float(s2) if s2 else 0.0

def _format_amount(val: float) -> str:
    return f"{val:.2f}"

def _sanitize_val(v: Optional[str]) -> Optional[str]:
    """Replace newlines with spaces (collapse to single spaces) and replace '|' with '/'.
    Keep text as-is otherwise (for display).
    """
    if v is None:
        return None
    s = str(v).replace("\r", " ").replace("\n", " ")
    s = " ".join(s.split())
    s = s.replace("|", "/")
    s = s.strip()
    return s if s else None

def _norm_acc(v: Optional[str]) -> str:
    """Normalize account numbers/IBANs for comparison (remove spaces, upper-case)."""
    return re.sub(r"\s+", "", (v or "")).upper()

def _get_owner_accounts(self) -> Set[str]:
    """Owner account sourced **only from context**:
       - env.context['journal_id'], or
       - env.context['active_model']=='account.journal' and env.context['active_id']
       If none found → empty set.
    """
    owners: Set[str] = set()
    ctx = self.env.context or {}
    jid = ctx.get("journal_id")
    if not jid and ctx.get("active_model") == "account.journal":
        jid = ctx.get("active_id")
    try:
        if jid:
            j = self.env["account.journal"].browse(int(jid))
            if j and getattr(j, "bank_account_id", False):
                owners.add(_norm_acc(j.bank_account_id.acc_number))
    except Exception:
        pass
    owners.discard("")
    _logger.debug("BA sheet: owner accounts from context → %s", sorted(list(owners)) or ["<none>"])
    return owners

def _choose_partner_name(self, row: Dict[str, object]) -> Optional[str]:
    """Strict rule with context-only owner account(s):
    - Require BOTH payer/payee account numbers present.
    - If exactly one of them belongs to owner_accounts, return the OTHER side's name.
    - Else: return None.
    """
    owner_accounts = _get_owner_accounts(self)
    if not owner_accounts:
        return None

    payer_name = _sanitize_val(row.get("payer name"))
    payee_name = _sanitize_val(row.get("payee name"))
    payer_acc = _norm_acc(row.get("payer account"))
    payee_acc = _norm_acc(row.get("payee account"))
    if not payer_acc or not payee_acc:
        return None

    payer_is_owner = payer_acc in owner_accounts
    payee_is_owner = payee_acc in owner_accounts

    if payer_is_owner ^ payee_is_owner:  # exactly one matches
        return (payee_name if payer_is_owner else payer_name) or None

    return None

def _build_payment_ref(row: Dict[str, object], amount: float, op_date_iso: str, val_date_iso: str) -> str:
    """Build the fixed, parseable payment_ref including BOTH payer and payee as provided."""
    pieces = []
    direction = "IN" if amount >= 0 else "OUT"
    pieces.append(f"DIR={direction}")

    bt = _sanitize_val(row.get("booking text"))
    if bt:
        pieces.append(f"BT={bt}")

    pieces.append(f"OD={op_date_iso}")
    pieces.append(f"VD={val_date_iso}")
    pieces.append(f"CUR={_sanitize_val(row.get('currency')) or 'EUR'}")
    pieces.append(f"AMT={_format_amount(amount)}")

    payer_name = _sanitize_val(row.get("payer name"))
    payer_acc_raw = _sanitize_val(row.get("payer account"))
    payer_bc = _sanitize_val(row.get("payer bank code"))
    payee_name = _sanitize_val(row.get("payee name"))
    payee_acc_raw = _sanitize_val(row.get("payee account"))
    payee_bc = _sanitize_val(row.get("payee bank code"))

    if payer_name:
        pieces.append(f"PAYER={payer_name}")
    if payer_acc_raw:
        pieces.append(f"PAYER_ACC={payer_acc_raw}")
    if payer_bc:
        pieces.append(f"PAYER_BC={payer_bc}")

    if payee_name:
        pieces.append(f"PAYEE={payee_name}")
    if payee_acc_raw:
        pieces.append(f"PAYEE_ACC={payee_acc_raw}")
    if payee_bc:
        pieces.append(f"PAYEE_BC={payee_bc}")

    pt = _sanitize_val(row.get("purpose text"))
    if pt:
        pieces.append(f"PT={pt}")
    ref = _sanitize_val(row.get("reference")) or _sanitize_val(row.get("record number"))
    if ref:
        pieces.append(f"REF={ref}")
    rd = _sanitize_val(row.get("record data"))
    if rd is not None:
        pieces.append(f"RD={rd}")

    return " | ".join(pieces)

class AccountStatementImportBASheet(models.TransientModel):
    _inherit = "account.statement.import"

    def _read_excel_rows_strict(self, content: bytes, kind: str):
        if kind == "xlsx":
            try:
                from openpyxl import load_workbook
            except Exception as e:
                raise UserError(_("Missing python dependency 'openpyxl' to read XLSX: %s") % e)
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.worksheets[0]
            it = ws.iter_rows(values_only=True)
            try:
                headers = [str(h or "").strip() for h in next(it)]
            except StopIteration:
                raise UserError(_("Empty Excel file."))
            norm = [_norm(h) for h in headers]
            idx = {}
            for i, h in enumerate(norm):
                if h in ALL_HEADERS:
                    idx[h] = i
            missing = [h for h in REQUIRED_HEADERS if h not in idx]
            if missing:
                raise UserError(_("Missing required columns: %s") % ", ".join(missing))
            rows = []
            for row in it:
                if not any(v not in (None, "") for v in row):
                    continue
                rec: Dict[str, object] = {}
                for key, col in idx.items():
                    if col < len(row):
                        rec[key] = row[col]
                rows.append(rec)
            return rows, idx
        else:
            try:
                import xlrd  # xlrd<2.0
            except Exception as e:
                raise UserError(_("To read legacy .XLS files, install 'xlrd<2.0' or export as XLSX. Error: %s") % e)
            book = xlrd.open_workbook(file_contents=content)
            sheet = book.sheet_by_index(0)
            if sheet.nrows == 0:
                raise UserError(_("Empty Excel file."))
            headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
            norm = [_norm(h) for h in headers]
            idx = {}
            for i, h in enumerate(norm):
                if h in ALL_HEADERS:
                    idx[h] = i
            missing = [h for h in REQUIRED_HEADERS if h not in idx]
            if missing:
                raise UserError(_("Missing required columns: %s") % ", ".join(missing))
            rows = []
            for r in range(1, sheet.nrows):
                if not any((str(sheet.cell_value(r, c)) if sheet.cell_value(r, c) is not None else "").strip() for c in range(sheet.ncols)):
                    continue
                rec: Dict[str, object] = {}
                for key, col in idx.items():
                    val = sheet.cell_value(r, col)
                    if key in ("operation date", "value date") and sheet.cell_type(r, col) == 3:
                        val = xlrd.xldate_as_datetime(val, book.datemode)
                    rec[key] = val
                rows.append(rec)
            return rows, idx

    def _parse_file(self, data_file):
        _logger.debug("BA sheet: start parsing")
        content = data_file if isinstance(data_file, bytes) else base64.b64decode(data_file)
        kind = _excel_kind(content)
        if kind is None:
            _logger.debug("BA sheet: not xls/xlsx -> passing to super()")
            return super()._parse_file(data_file)

        rows, header_idx = self._read_excel_rows_strict(content, kind)
        _logger.debug("BA sheet: read %d data rows from Excel.", len(rows))

        txs = []
        first_date = None
        last_date = None

        for idx, r in enumerate(rows, start=1):
            currency = str(r.get("currency") or "").strip().upper()
            if currency not in ("EUR", "€"):
                raise UserError(_("Non-EUR row detected (row %s): %s") % (idx, currency))

            amount = float(_parse_number(r.get("amount")))
            od_iso = _to_iso_date(r.get("operation date"))
            vd_iso = _to_iso_date(r.get("value date"))
            if od_iso:
                if first_date is None or od_iso < first_date:
                    first_date = od_iso
                if last_date is None or od_iso > last_date:
                    last_date = od_iso

            partner_name = _choose_partner_name(self, r)
            payref = _build_payment_ref(r, amount, od_iso, vd_iso)

            # unique id: date + amount + first 32 chars of booking text
            bt = _sanitize_val(r.get("booking text")) or ""
            bt_prefix = bt[:32]
            uid_seed = f"{od_iso}|{amount:.2f}|{bt_prefix}"
            import hashlib
            unique_import_id = hashlib.sha1(uid_seed.encode("utf-8")).hexdigest()

            tx = {
                "date": od_iso or fields.Date.today().isoformat(),
                "payment_ref": payref or _("Bank transaction"),
                "amount": amount,
                "unique_import_id": unique_import_id,
            }
            if partner_name:
                tx["partner_name"] = partner_name

            txs.append(tx)

        if not txs:
            raise UserError(_("No transactions found after validation."))

        # Sort lines ASC by date (and UID for stability)
        txs.sort(key=lambda t: (t["date"], t["unique_import_id"]))

        # Statement dates and name
        stmt_date = last_date or fields.Date.today().isoformat()
        if first_date and last_date and first_date != last_date:
            stmt_name = _("Bank Austria import %s..%s (EUR)") % (first_date, last_date)
        else:
            sd = first_date or stmt_date
            stmt_name = _("Bank Austria import %s (EUR)") % sd
        _logger.debug("BA sheet: tx count=%d; sorted ASC; date range=%s..%s", len(txs), first_date, last_date)

        stmt_vals = {
            "date": stmt_date,
            "transactions": txs,
            "name": stmt_name,
        }

        payload = [("EUR", None, [stmt_vals])]
        _logger.debug("BA sheet: returning 3-tuple payload")
        return payload

